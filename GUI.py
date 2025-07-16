from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

app = Flask(__name__)

archivo = "ventas.xlsx"

if not os.path.exists(archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha", "Mesa", "Producto", "Cantidad", "Precio Unitario", "Total"])
    wb.save(archivo)

precios = {
    "Local": {"Hamburguesa": 60, "Pizza": 80, "Tacos": 30},
    "Rappi": {"Hamburguesa": 70, "Pizza": 90, "Tacos": 35},
    "Uber": {"Hamburguesa": 75, "Pizza": 95, "Tacos": 40},
}

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/calcular_total", methods=["POST"])
def calcular_total():
    data = request.json
    mesa = data["mesa"]
    producto = data["producto"]
    cantidad = int(data["cantidad"])
    precio = precios[mesa][producto]
    total = precio * cantidad
    return jsonify({"total": total})

@app.route("/registrar_venta", methods=["POST"])
def registrar_venta():
    data = request.json
    mesa = data["mesa"]
    producto = data["producto"]
    cantidad = int(data["cantidad"])
    precio_unitario = precios[mesa][producto]
    total = precio_unitario * cantidad
    fecha = datetime.now().strftime("%Y-%m-%d")

    wb = load_workbook(archivo)
    ws = wb["Ventas"]
    ws.append([fecha, mesa, producto, cantidad, precio_unitario, total])
    wb.save(archivo)
    return jsonify({"mensaje": "Venta registrada"})

@app.route("/reportes", methods=["POST"])
def reportes():
    fecha = request.json["fecha"]
    try:
        wb = load_workbook(archivo)
        ws = wb["Ventas"]
        productos = {}
        totales_por_mesa = {"Local": 0, "Rappi": 0, "Uber": 0}
        total_dia = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == fecha:
                mesa, producto, cantidad, total = row[1], row[2], row[3], row[5]
                productos[producto] = productos.get(producto, 0) + cantidad
                totales_por_mesa[mesa] += total
                total_dia += total

        return jsonify({
            "productos": productos,
            "mesas": totales_por_mesa,
            "total": total_dia
        })
    except:
        return jsonify({"error": "Error en archivo o fecha"}), 400

@app.route("/cancelar", methods=["POST"])
def cancelar():
    fila = int(request.json["fila"])
    try:
        wb = load_workbook(archivo)
        ws = wb["Ventas"]
        ws.delete_rows(fila)
        wb.save(archivo)
        return jsonify({"mensaje": f"Fila {fila} eliminada"})
    except:
        return jsonify({"error": "Error al eliminar"}), 400

if __name__== "_main_":
    app.run(debug=True)